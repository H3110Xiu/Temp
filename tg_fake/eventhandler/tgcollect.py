# -*- coding:utf-8 -*-

from PyQt5.QtCore import pyqtSignal, QObject, QDateTime
from PyQt5.QtWidgets import QFileDialog, QMessageBox, QLabel
from PyQt5.QtGui import QTextCursor
from PyQt5 import QtWidgets

from telethon import TelegramClient
from telethon import functions, types
import asyncio
import re

import time
import requests
from urllib.parse import quote, unquote
import random
import os
import openpyxl
import shutil
from threading import Thread
import configparser

from ui.collectui import Ui_Form


# 定义信号
class MySignal(QObject):
    # 输出日志信号
    print_log = pyqtSignal(str)
    # 在线程中使用信息框信号，防止界面卡死
    show_message_box = pyqtSignal(str, str)
    # 在label显示信息
    show_info = pyqtSignal(QLabel, str)


class CollectWindow(QtWidgets.QWidget):
    def __init__(self, parent=None):
        super(CollectWindow, self).__init__(parent)
        self.ui = Ui_Form()
        self.ui.setupUi(self)

        # 定义全局变量
        self.api_id = 2678580
        self.api_hash = "34b88c2e20ec72c0d722138200615fb8"
        self.proxy_type = "socks5"
        self.collect_type = "线路1"
        self.proxy_link = ""
        self.keyword = ""
        self.group_file_path = ""
        self.session_path = ""
        self.group_list = []
        self.group_count = 0
        self.used_sessions_dir = ""
        self.proxy_delay = 5
        self.session_list = []
        self.keyword_list = []
        self.run_flag = True
        self.save_type = "txt"

        # 读取配置文件
        self.get_values("conf.ini")

        # 创建信号，绑定信号事件
        self.ms = MySignal()
        self.ms.print_log.connect(self.output_log)
        self.ms.show_message_box.connect(self.show_message_box)
        self.ms.show_info.connect(self.show_info)

        # 绑定界面控件事件
        self.ui.pushButton_url.clicked.connect(self.select_group_file_path)
        self.ui.pushButton_session.clicked.connect(self.select_session_path)
        self.ui.pushButton_ip.clicked.connect(self.test_proxy)
        # self.ui.pushButton_start.clicked.connect(self.start_collect)
        self.ui.pushButton_start.clicked.connect(self.start_collect)
        self.ui.pushButton_stop.clicked.connect(self.stop_collect)

    # 读配置文件
    def get_values(self, path):
        if os.path.exists(path):
            config = configparser.ConfigParser()
            config.read(path, encoding="utf-8-sig")

            group_path = config.get("collect_file", "group_path")
            session_path = config.get("collect_file", "session_path")

            ip_link = config.get("collect_proxy", "ip_link")

            collect_keyword = config.get("collect_collect", "collect_keyword")

            self.ui.lineEdit_url.setText(group_path)
            self.ui.lineEdit_session.setText(session_path)

            self.ui.lineEdit_ip.setText(ip_link)

            self.ui.lineEdit_keyword.setText(collect_keyword)
        else:
            return

    # 写配置文件
    def set_values(self, path):
        if os.path.exists(path):
            config = configparser.ConfigParser()
            config.read(path, encoding="utf-8-sig")

            group_path = self.ui.lineEdit_url.text().strip()
            session_path = self.ui.lineEdit_session.text().strip()

            ip_link = self.ui.lineEdit_ip.text().strip()

            collect_keyword = self.ui.lineEdit_keyword.text().strip()

            config.set("collect_file", "group_path", group_path)
            config.set("collect_file", "session_path", session_path)

            config.set("collect_proxy", "ip_link", ip_link)

            config.set("collect_collect", "collect_keyword", collect_keyword)

            config.write(open(path, "w", encoding="utf-8"))
        else:
            return

    # 输出日志信号
    def output_log(self, str):
        date_time = QDateTime.currentDateTime().toString("hh:mm:ss")
        self.ui.textEdit_log.append(date_time + " " + str)
        self.ui.textEdit_log.ensureCursorVisible()
        cursor = self.ui.textEdit_log.textCursor()
        cursor.movePosition(QTextCursor.End)

    # 显示信息框
    def show_message_box(self, type, str):
        if type == "警告":
            QMessageBox.warning(self, type, str)
        else:
            QMessageBox.information(self, type, str)

    # 显示信息
    def show_info(self, label, str):
        label.setText(str)

    # 选择保存群组链接的文件夹
    def select_group_file_path(self):
        file_name = QFileDialog.getExistingDirectory(None, "选择群组链接保存文件夹", "./")
        self.ui.lineEdit_url.setText(file_name)
        if not file_name:
            return
        self.ms.print_log.emit("已选择群组链接保存文件夹:" + file_name)

    # 选择session文件
    def select_session_path(self):
        file_name = QFileDialog.getExistingDirectory(None, "选择session文件夹", "./")
        self.ui.lineEdit_session.setText(file_name)
        if not file_name:
            return
        self.ms.print_log.emit("已选择session文件夹:" + file_name)

    # 测试代理线程
    def test_proxy_thread(self):
        url = self.ui.lineEdit_ip.text().strip()
        if url == "":
            self.ms.show_message_box.emit("提示", "请输入代理ip链接")
        else:
            try:
                proxy_ip = requests.get(url).text
                self.ms.show_message_box.emit("提示", proxy_ip)
            except:
                self.ms.show_message_box.emit("提示", "获取ip失败")

    # 测试代理
    def test_proxy(self):
        thread = Thread(target=self.test_proxy_thread)
        thread.setDaemon(True)
        thread.start()

    # 获取代理ip
    def get_proxy_ip(self):
        self.ms.print_log.emit("开始获取代理IP")
        try:
            rtn = requests.get(self.proxy_link).text
            self.ms.print_log.emit("获取到代理IP:" + rtn)
            self.ms.print_log.emit("完成一次IP获取,等待" + str(self.proxy_delay) + "秒")
            time.sleep(self.proxy_delay)
            return rtn
        except Exception as e:
            print(e)
            self.ms.print_log.emit("获取代理IP失败")
            self.ms.print_log.emit("完成一次IP获取,等待" + str(self.proxy_delay) + "秒")
            time.sleep(self.proxy_delay)
            return

    # 通过群组链接获取群组信息Excel
    def save_group_to_excel(self, file_path, group_list):
        self.group_count = 0
        work_book = openpyxl.Workbook()
        sheet = work_book.active
        sheet.cell(1, 1, "群链接")

        index = 2
        for group in group_list:
            if group["类型"]:
                continue
            else:
                self.group_count += 1
                sheet.cell(index, 1, str(group["群链接"]))
            index += 1
        self.ms.show_info.emit(self.ui.label_collect_count, str(self.group_count))
        work_book.save(file_path)

    # 通过群组链接获取群组信息TXT
    def save_group_to_txt(self, file_path, group_list):
        with open(file_path, "w+", encoding="utf-8") as file:
            file.write("群链接" + "\n")
            for group in group_list:
                if group["类型"]:
                    continue
                else:
                    file.write(str(group["群链接"]) + "\n")

    # 保存群组链接到文件
    def save_group_file(self, file_path, group_list):
        self.ms.print_log.emit("-----开始导出群组链接文件-----")
        with open(file_path, "a+") as file:
            for group in group_list:
                file.write(group + "\n")
        self.ms.print_log.emit("-----导出群组链接文件完成-----")
        self.ms.print_log.emit("群组链接保存在" + self.group_file_path + "目录下")

    # 获取群组详细信息，并保存为excel
    async def get_group_info(self, session, url_list, key):
        # 如果列表不为空
        if url_list:
            self.ms.print_log.emit("-----开始获取群组详细信息-----")
            while self.run_flag:
                ip = self.get_proxy_ip()
                group_info = []
                try:
                    self.ms.print_log.emit("正在登录......")
                    async with TelegramClient(session, self.api_id, self.api_hash, proxy=(
                            self.proxy_type, ip.split(":")[0], int(ip.split(":")[1])), timeout=5) as client:
                        self.ms.print_log.emit("登录成功")
                        for url in url_list:
                            try:
                                result = await client(functions.channels.GetFullChannelRequest(
                                    channel=url
                                ))
                                group_info.append({"群ID": result.full_chat.id, "群组名": result.chats[0].title, "群链接": url,
                                                   "类型": result.chats[0].broadcast, "群成员数": result.full_chat.participants_count})
                                self.ms.print_log.emit(url + "---获取群组详细信息成功")
                                self.ms.print_log.emit("-----开始保存群组信息-----")
                                if self.save_type == "excel":
                                    new_path_excel = self.group_file_path + "/" + key + ".xlsx"
                                    self.save_group_to_excel(new_path_excel, group_info)
                                if self.save_type == "txt":
                                    new_path_txt = self.group_file_path + "/" + key + ".txt"
                                    self.save_group_to_txt(new_path_txt, group_info)
                                self.ms.print_log.emit("-----保存群组信息完成-----")
                            except Exception as e:
                                print("1----", e)
                                self.ms.print_log.emit(url + "---获取群组详细信息失败")
                            if not self.run_flag:
                                return
                        break
                except:
                    self.ms.print_log.emit("登录账号失败，获取群组信息失败")
                    continue

            self.ms.print_log.emit("-----获取群组详细信息完成-----")

            # 移动已使用过的session
            await client.disconnect()
            session_name = session.split("/")[-1]
            date_time = QDateTime.currentDateTime().toString("yyyy.MM.dd")
            root_path = os.path.dirname(self.session_path)
            if not root_path:
                self.used_sessions_dir = date_time + "已经采集过群组的sessions"
            else:
                self.used_sessions_dir = root_path + "/" + date_time + "已经采集过群组的sessions"
            if not os.path.exists(self.used_sessions_dir):
                self.ms.print_log.emit("创建<已经采集过群组的sessions>目录")
                os.mkdir(self.used_sessions_dir)
            shutil.move(session, self.used_sessions_dir + "/" + session_name)
        else:
            return

    # 获取session列表
    def get_session_list(self, sessions_path):
        session_list = []
        for file in os.listdir(sessions_path):
            if file.endswith(".session"):
                session_list.append(sessions_path + "/" + file)
        return session_list

    # 获取关键词
    def get_keyword_list(self):
        keyword_list = []
        if "," in self.keyword:
            keyword_list = [key for key in self.keyword.split(",") if key]
            return keyword_list
        if "，" in self.keyword:
            keyword_list = [key for key in self.keyword.split("，") if key]
            return keyword_list
        keyword_list.append(self.keyword)
        return keyword_list

    # 采集线程
    def collect_thread(self):
        self.ms.print_log.emit("-----开始采集群组-----")

        # 外核搜索
        if self.collect_type == "线路1":
            async def md_search(session, key):
                # 获取cse_token
                cse_tok = ""
                token_count = 0
                while True:
                    token_count += 1
                    token_ip = self.get_proxy_ip()
                    proxies = {
                        "http": self.proxy_type + "://" + token_ip,
                        "https": self.proxy_type + "://" + token_ip
                    }
                    self.ms.print_log.emit("正在连接......")
                    data = '"cse_token": "(.*?)",'
                    token_url = "https://cse.google.com/cse.js?cx=006249643689853114236:meozern20ky"
                    if token_count > 200:
                        self.ms.print_log.emit("由于服务器原因连接失败")
                        return
                    try:
                        token_ret = requests.get(token_url, proxies=proxies).text
                    except:
                        continue
                    cse_tok = re.findall(data, token_ret)[0]
                    break


                start = 0
                while self.run_flag:
                    # 获取代理ip
                    ip = self.get_proxy_ip()
                    proxies = {
                        "http": self.proxy_type + "://" + ip,
                        "https": self.proxy_type + "://" + ip
                    }

                    self.ms.print_log.emit("正在采集群组......")

                    url = "https://cse.google.com/cse/element/v1?rsz=filtered_cse&num=20&hl=zh-CN&source=gcsc" \
                          "&gss=.io&start=" + str(start) + "&cselibv=323d4b81541ddb5b&cx=006249643689853114236:meozern20ky" \
                          "&q=" + quote(key) + "%20more%3A%E7%BE%A4%E7%BB%84&safe=active" \
                          "&cse_tok=" + cse_tok + "&sort=&exp=csqr,cc" \
                          "&callback=google.search.cse.api" + str(random.randint(1000, 9999))

                    ret = ""
                    try:
                        ret = requests.get(url, proxies=proxies, timeout=10).text
                    except Exception as e:
                        print("2----", e)
                        self.ms.print_log.emit("连接服务器失败，获取群组链接失败")
                        continue
                    if not ret.startswith("/*O_o*/") or ("results" not in ret):
                        self.ms.print_log.emit("-----群组采集结束-----")
                        break
                    self.ms.print_log.emit("采集到群组链接:")
                    self.ms.print_log.emit("正在进行群组链接过滤...")
                    glo = {
                        "true": 1,
                        "false": 0
                    }
                    try:
                        dic_data = eval(unquote(ret[33:-1]), glo)
                        for data in dic_data["results"]:
                            count = data["url"].count("/")
                            if (count == 3) and (data["url"] not in self.group_list):
                                self.ms.print_log.emit(data["url"])
                                self.group_list.append(data["url"])
                        start += 20
                    except:
                        self.ms.print_log.emit("获取格式错误")
                        continue

                if self.run_flag:
                    # 采集结束，获取群组详细信息并写出群组链接到文件
                    await self.get_group_info(session, self.group_list, key)
                    self.ms.show_info.emit(self.ui.label_collect_count, str(self.group_count))
            if len(self.keyword_list) < len(self.session_list):
                session_index = 0
                for keyword in self.keyword_list:
                    self.group_list = []
                    md_loop = asyncio.new_event_loop()
                    asyncio.set_event_loop(md_loop)
                    md_loop.run_until_complete(md_search(self.session_list[session_index], keyword))
                    session_index += 1
                self.ms.show_message_box.emit("提示", "群组采集完成")
            else:
                keyword_index = 0
                for session in self.session_list:
                    self.group_list = []
                    md_loop = asyncio.new_event_loop()
                    asyncio.set_event_loop(md_loop)
                    md_loop.run_until_complete(md_search(session, self.keyword_list[keyword_index]))
                    keyword_index += 1
                self.ms.show_message_box.emit("提示", "群组采集完成")

        # 内核搜索
        if self.collect_type == "线路2":
            async def collect(session, key):
                while self.run_flag:
                    ip = self.get_proxy_ip()
                    try:
                        self.ms.print_log.emit("正在登录......")
                        async with TelegramClient(session, self.api_id, self.api_hash,
                                                  proxy=(
                                                  self.proxy_type, ip.split(":")[0], int(ip.split(":")[1])), timeout=5) as client:
                            self.ms.print_log.emit("登录成功")
                            try:
                                ret = await client.send_message("hao1234bot", key)
                            except Exception as e:
                                self.ms.print_log.emit("采集群组链接失败")
                                await client.disconnect()
                                return
                            time.sleep(5)
                            self.ms.print_log.emit("正在采集群组......")
                            while self.run_flag:
                                try:
                                    messages = await client.get_messages("hao1234bot")
                                except Exception as e:
                                    self.ms.print_log.emit("采集群组链接失败")
                                    await client.disconnect()
                                    return
                                pattern = re.compile("[(](.*?)[)]", re.S)
                                temp_list = re.findall(pattern, messages[0].text)
                                if len(temp_list) >= 2:
                                    self.ms.print_log.emit("采集到群组链接:")
                                    for url in temp_list[1:]:
                                        if url.count("/") < 3:
                                            continue
                                        self.ms.print_log.emit(url)
                                        self.group_list.append(url)
                                else:
                                    self.ms.print_log.emit("未采集到群组链接")
                                    return
                                try:
                                    await messages[0].click(21)
                                except:
                                    break
                            self.ms.print_log.emit("-----群组采集结束-----")
                            await client.disconnect()

                        await self.get_group_info(session, self.group_list, key)
                        self.ms.show_info.emit(self.ui.label_collect_count, str(self.group_count))
                        return
                    except Exception as e:
                        print("3----", e)
                        self.ms.print_log.emit("登录账号失败，获取群组链接失败")
            if len(self.keyword_list) < len(self.session_list):
                session_index = 0
                for keyword in self.keyword_list:
                    self.group_list = []
                    collect_loop = asyncio.new_event_loop()
                    asyncio.set_event_loop(collect_loop)
                    collect_loop.run_until_complete(collect(self.session_list[session_index], keyword))
                    session_index += 1
                self.ms.show_message_box.emit("提示", "群组采集完成")
            else:
                keyword_index = 0
                for session in self.session_list:
                    self.group_list = []
                    collect_loop = asyncio.new_event_loop()
                    asyncio.set_event_loop(collect_loop)
                    collect_loop.run_until_complete(collect(session, self.keyword_list[keyword_index]))
                    keyword_index += 1
                self.ms.show_message_box.emit("提示", "群组采集完成")

    # 开始采集
    def start_collect(self):
        self.ui.tabWidget.setCurrentIndex(1)
        # 写出配置文件
        self.set_values("conf.ini")
        # 初始化变量
        self.proxy_link = ""
        self.keyword = ""
        self.group_file_path = ""
        self.session_path = ""
        self.group_list = []
        self.group_count = 0
        self.used_sessions_dir = ""
        self.session_list = []
        self.keyword_list = []
        self.run_flag = True

        # 检查信息填写是否完整

        # 检查群组链接保存文件夹是否选择
        self.group_file_path = self.ui.lineEdit_url.text().strip()
        if not self.group_file_path:
            self.ms.show_message_box.emit("警告", "未选择群组链接保存文件夹!")
            return
        self.ms.print_log.emit("选择的群组链接保存文件夹为:" + self.group_file_path)

        # 检查session文件是否选择
        self.session_path = self.ui.lineEdit_session.text().strip()
        if not self.session_path:
            self.ms.show_message_box.emit("警告", "未选择session文件夹!")
            return
        self.ms.print_log.emit("选择的session文件夹为:" + self.session_path)

        # 获取session列表
        try:
            self.session_list = self.get_session_list(self.session_path)
        except:
            self.ms.print_log.emit("读取session文件失败，请检查后重新尝试")
            return

        # 获取保存格式
        if self.ui.radioButton_txt.isChecked():
            self.save_type = "txt"
        if self.ui.radioButton_excel.isChecked():
            self.save_type = "excel"

        # 检查代理链接是否填写
        self.proxy_link = self.ui.lineEdit_ip.text().strip()
        if not self.proxy_link:
            self.ms.show_message_box.emit("警告", "未指定代理ip链接!")
            return

        # 设置采集方式
        self.collect_type = self.ui.comboBox_collect_type.currentText()

        # 检查关键词是否填写
        self.keyword = self.ui.lineEdit_keyword.text().strip()
        if not self.keyword:
            self.ms.show_message_box.emit("警告", "未指定搜索关键词!")
            return
        self.ms.print_log.emit("指定的搜索关键词为:" + "<" + self.keyword + ">")

        try:
            self.keyword_list = self.get_keyword_list()
        except:
            self.ms.print_log.emit("获取关键词失败,请检查后重新尝试!")
            return

        # 开启采集线程
        thread = Thread(target=self.collect_thread)
        thread.setDaemon(True)
        thread.start()

    # 停止采集
    def stop_collect(self):
        self.ms.print_log.emit("正在停止采集...")
        self.run_flag = False

