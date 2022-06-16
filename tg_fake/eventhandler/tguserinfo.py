from PyQt5.QtCore import pyqtSignal, QObject, QDateTime
from PyQt5.QtWidgets import QFileDialog, QMessageBox, QLabel
from PyQt5.QtGui import QTextCursor
from PyQt5 import QtWidgets
from telethon import TelegramClient
from telethon.tl.types import UserStatusOnline, UserStatusOffline
import time
import asyncio
import requests
import os
import shutil
import configparser
from datetime import datetime, timezone, timedelta
import openpyxl
from threading import Thread, Lock
from concurrent.futures import ThreadPoolExecutor, as_completed

from ui.getuserinfo import Ui_Form


# 定义信号
class MySignal(QObject):
    # 输出日志信号
    print_log = pyqtSignal(str)
    # 线程中使用信息框信号，防止界面卡死
    show_message_box = pyqtSignal(str, str)
    # 在label显示群组和用户信息
    show_info = pyqtSignal(QLabel, str)
    # 更新进度条
    update_process_bar = pyqtSignal(int)


class UserWindow(QtWidgets.QWidget):
    def __init__(self, parent = None):
        super(UserWindow, self).__init__(parent)
        self.ui = Ui_Form()
        self.ui.setupUi(self)

        # 定义全局变量
        self.api_id = 2678580
        self.api_hash = "34b88c2e20ec72c0d722138200615fb8"
        self.send_type = 0
        self.proxy_type = "socks5"
        self.session_path = ""
        self.user_info_path = ""
        self.proxy_link = ""
        self.obtain_group_count = 5
        self.obtain_delay = 5
        self.thread_count = 0
        self.group_count = 0
        self.group_file_path = ""
        self.all_group_list = []
        self.group_list_index = 0
        self.group_list_index_lock = Lock()
        self.process_bar_value = 0
        self.process_bar_value_lock = Lock()
        self.proxy_delay = 5
        self.login_time = 0
        self.proxy_ip_lock = Lock()
        self.save_type = ""

        # 读取配置文件
        self.get_values("conf.ini")

        # 创建信号，绑定信号事件
        self.ms = MySignal()
        self.ms.print_log.connect(self.output_log)
        self.ms.show_message_box.connect(self.show_message_box)
        self.ms.show_info.connect(self.show_info)
        self.ms.update_process_bar.connect(self.update_process_bar)

        # 绑定界面控件事件
        self.ui.pushButton_group.clicked.connect(self.select_group_file)
        self.ui.pushButton_sessions.clicked.connect(self.select_sessions)
        self.ui.pushButton_user_info.clicked.connect(self.select_user_info)
        self.ui.pushButton_test_ip.clicked.connect(self.test_ip)
        self.ui.pushButton_start.clicked.connect(self.start_obtain)
        self.ui.pushButton_stop.clicked.connect(self.stop_obtain)
        # self.ui.pushButton_output_userinfo.clicked.connect(self.save_user_info)

    # 读配置文件
    def get_values(self, path):
        if os.path.exists(path):
            config = configparser.ConfigParser()
            config.read(path, encoding="utf-8-sig")

            group_path = config.get("user_info_file", "group_path")
            session_path = config.get("user_info_file", "session_path")
            info_path = config.get("user_info_file", "info_path")
            save_type = config.get("user_info_file", "save_type")

            ip_link = config.get("user_info_proxy", "ip_link")

            thread_amount = config.get("user_info_software", "thread_amount")
            login_time = config.get("user_info_software", "login_time")

            self.ui.lineEdit_group.setText(group_path)
            self.ui.lineEdit_sessions.setText(session_path)
            self.ui.lineEdit_user_info.setText(info_path)
            if save_type == "txt":
                self.ui.radioButton_txt.setChecked(True)
            if save_type == "excel":
                self.ui.radioButton_excel.setChecked(True)

            self.ui.comboBox_thread.setCurrentText(thread_amount)
            self.ui.comboBox_login_time.setCurrentText(login_time)
        else:
            return

    # 写配置文件
    def set_values(self, path):
        if os.path.exists(path):
            config = configparser.ConfigParser()
            config.read(path, encoding="utf-8-sig")

            group_path = self.ui.lineEdit_group.text().strip()
            session_path = self.ui.lineEdit_sessions.text().strip()
            info_path = self.ui.lineEdit_user_info.text().strip()
            save_type = ""
            if self.ui.radioButton_txt.isChecked():
                save_type = "txt"
            if self.ui.radioButton_excel.isChecked():
                save_type = "excel"

            ip_link = self.ui.lineEdit_ip.text().strip()

            thread_amount = self.ui.comboBox_thread.currentText().strip()
            login_time = self.ui.comboBox_login_time.currentText().strip()

            config.set("user_info_file", "group_path", group_path)
            config.set("user_info_file", "session_path", session_path)
            config.set("user_info_file", "info_path", info_path)
            config.set("user_info_file", "save_type", save_type)

            config.set("user_info_proxy", "ip_link", ip_link)

            config.set("user_info_software", "thread_amount", thread_amount)
            config.set("user_info_software", "login_time", login_time)

            config.write(open(path, "w", encoding="utf-8"))
        else:
            return

    # 输出日志信号
    def output_log(self, str):
        date_time = QDateTime.currentDateTime().toString("hh:mm:ss")
        self.ui.textEdit_log.append(date_time + " " + str)
        self.ui.textEdit_log.ensureCursorVisible()
        cursor = self.ui.textEdit_log.textCursor()
        cursor.movePosition(QTextCursor.End)

    # 显示信息框
    def show_message_box(self, type, str):
        if type == "警告":
            QMessageBox.warning(self, type, str)
        else:
            QMessageBox.information(self, type, str)

    # 显示信息
    def show_info(self, label, str):
        label.setText(str)

    # 更新进度条
    def update_process_bar(self, value):
        self.ui.progressBar.setValue(value)

    # 选择群组链接文件
    def select_group_file(self):
        file_name, file_type = QFileDialog.getOpenFileName(None, "选择群组文件", "./",
                                                            "文本文档(*.txt)")
        self.ui.lineEdit_group.setText(file_name)
        if not file_name:
            return
        self.ms.print_log.emit("已选择群组文件:" + file_name)

    # 选择sessions文件夹
    def select_sessions(self):
        file_name = QFileDialog.getExistingDirectory(None, "选择sessions文件夹", "./")
        self.ui.lineEdit_sessions.setText(file_name)
        if not file_name:
            return
        self.ms.print_log.emit("已选择sessions文件夹:" + file_name)

    # 选择user_info文件夹
    def select_user_info(self):
        file_name = QFileDialog.getExistingDirectory(None, "选择用户信息保存文件夹", "./")
        self.ui.lineEdit_user_info.setText(file_name)
        if not file_name:
            return
        self.ms.print_log.emit("已选择用户信息保存文件夹:" + file_name)

    # 测试代理IP线程
    def test_ip_thread(self):
        url = self.ui.lineEdit_ip.text().strip()
        if url == "":
            self.ms.show_message_box.emit("提示", "请输入代理ip链接")
        else:
            try:
                proxy_ip = requests.get(url).text
                self.ms.show_message_box.emit("提示", proxy_ip)
            except:
                self.ms.show_message_box.emit("提示", "获取ip失败")

    # 测试代理IP
    def test_ip(self):
        thread = Thread(target=self.test_ip_thread)
        thread.setDaemon(True)
        thread.start()

    # 获取代理ip
    def get_proxy_ip(self):
        self.ms.print_log.emit("开始获取代理IP")
        try:
            rtn = requests.get(self.proxy_link).text
            self.ms.print_log.emit("获取到代理IP:" + rtn)
            self.ms.print_log.emit("完成一次IP获取,等待" + str(self.proxy_delay) + "秒")
            time.sleep(self.proxy_delay)
            proxy_ip = rtn.split(":")
            return proxy_ip
        except:
            self.ms.print_log.emit("获取代理IP失败")
            self.ms.print_log.emit("完成一次IP获取,等待" + str(self.proxy_delay) + "秒")
            time.sleep(self.proxy_delay)
            return

    # 获取群组链接
    def get_group_list(self, path):
        group_list = []
        with open(path, "r") as file:
            for line in file.readlines():
                group_list.append(line.strip("\n"))

        return group_list

    # 获取session列表
    def get_session_list(self, sessions_path):
        session_list = []
        for file in os.listdir(sessions_path):
            session_list.append(sessions_path + "/" + file)
        return session_list

    # 保存群组成员信息到Excel
    def save_user_info_to_excel(self, file_path, user_list):
        self.group_count = 0
        work_book = openpyxl.Workbook()
        sheet = work_book.active
        sheet.cell(1, 1, "username")

        index = 2
        for user in user_list:
            sheet.cell(index, 1, str(user))
            index += 1
        work_book.save(file_path)

    # 保存群组成员信息到TXT
    def save_user_info_to_txt(self, file_path, user_list):
        with open(file_path, "w+", encoding="utf-8") as file:
            file.write("username" + "\n")
            for id_username in user_list:
                file.write(str(id_username) + "\n")

    # 获取群组成员信息
    async def obtain_user_info(self, session, ip, group_url):
        # 更新进度条
        self.process_bar_value_lock.acquire()
        self.process_bar_value += 1
        self.ms.update_process_bar.emit(self.process_bar_value)
        self.process_bar_value_lock.release()

        user_list = []
        file_name_txt = self.user_info_path + "/" + group_url.split("/")[-1] + ".txt"
        file_name_excel = self.user_info_path + "/" + group_url.split("/")[-1] + ".xlsx"

        try:
            async with TelegramClient(session, self.api_id, self.api_hash,
                                  proxy=(self.proxy_type, ip[0], int(ip[1]))) as client:
                try:
                    self.ms.print_log.emit(group_url + "-----开始获取群组成员信息")
                    self.ms.print_log.emit(group_url + "-----正在获取群组成员信息...")
                    users = await client.get_participants(group_url)
                except:
                    self.ms.print_log.emit(group_url + "-----群组无效")
                    self.ms.print_log.emit(group_url + "-----获取群组成员信息失败")
                    return
                for user in users:
                    # 获取未删除账号的信息
                    if user.status:
                        if self.login_time:
                            # 在线账号
                            if isinstance(user.status, UserStatusOnline):
                                if user.username:
                                    user_list.append(user.username)
                            # 在线时间在指定范围内
                            if isinstance(user.status, UserStatusOffline):
                                temp_time = timedelta(hours=self.login_time)
                                if (user.status.was_online + temp_time) >= datetime.now(timezone.utc):
                                    if user.username:
                                        user_list.append(user.username)
                        else:
                            if user.username or user.phone:
                                user_list.append(user.username)
            self.ms.print_log.emit(group_url + "-----获取群组成员信息成功")
            self.ms.print_log.emit(group_url + "-----开始保存群组成员信息")
            if self.save_type == "txt":
                self.save_user_info_to_txt(file_name_txt, user_list)
            if self.save_type == "excel":
                self.save_user_info_to_excel(file_name_excel, user_list)
            self.ms.print_log.emit(group_url + "-----保存群组成员信息完成")
        except:
            self.ms.print_log.emit(group_url + "-----登录账号失败")
            self.ms.print_log.emit(group_url + "-----获取群组成员信息失败")
            return

    # 一个账号获取指定个数群组成员信息线程
    def obtain_user_info_thread(self, session, group_list):
        session_name = session.split("/")[-1]
        self.ms.print_log.emit(session_name + "-----开始获取群组成员信息")
        # 获取待获取群组数
        group_to_obtain_list = []
        self.group_list_index_lock.acquire()
        if len(group_list) < self.obtain_group_count:
            self.obtain_group_count = len(group_list)
        if (self.group_list_index + self.obtain_group_count) > len(group_list):
            group_to_obtain_list = group_list[self.group_list_index:]
            self.group_list_index = len(group_list)
        else:
            group_to_obtain_list = group_list[self.group_list_index:self.group_list_index + self.obtain_group_count]
            self.group_list_index += self.obtain_group_count

        # group_to_obtain_list = group_list[self.group_list_index:self.group_list_index + self.obtain_group_count]
        # self.group_list_index += self.obtain_group_count

        # 删除群组文件已经使用群组链接
        new_path = self.group_file_path.replace(".txt", "_未获取群组成员信息.txt")
        with open(new_path, "w+") as file:
            for url in self.all_group_list[self.group_list_index:]:
                file.write(url + "\n")

        self.group_list_index_lock.release()

        for count in range(len(group_to_obtain_list)):
            # 获取代理ip
            self.proxy_ip_lock.acquire()
            ip = self.get_proxy_ip()
            self.proxy_ip_lock.release()
            group_url = group_to_obtain_list[count]
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            loop.run_until_complete(self.obtain_user_info(session, ip, group_url))
            if count < self.obtain_group_count - 1:
                self.ms.print_log.emit("完成一个群组的成员信息获取，休息" + str(self.obtain_delay) + "秒")
                time.sleep(self.obtain_delay)

        self.ms.print_log.emit(session_name + "-----获取群组成员信息完成")
        shutil.move(session, self.used_sessions_dir + "/" + session_name)
        # self.ms.print_log.emit("-----发送群组消息任务完成-----")
        # self.ms.show_message_box.emit("提示", "发送群组消息完成")

    # 子线程，开启线程池
    def start_obtain_thread(self, group_list, session_list):
        # 创建文件夹保存已经使用过的sessions
        date_time = QDateTime.currentDateTime().toString("yyyy.MM.dd")
        root_path = os.path.dirname(self.session_path)
        if not root_path:
            self.used_sessions_dir = date_time + "已经获取过成员信息的sessions"
        else:
            self.used_sessions_dir = root_path + "/" + date_time + "已经获取过成员信息的sessions"
        if not os.path.exists(self.used_sessions_dir):
            self.ms.print_log.emit("创建<已经获取过成员信息的sessions>目录")
            os.mkdir(self.used_sessions_dir)

        with ThreadPoolExecutor(max_workers=self.thread_count) as pool:
            self.future_list = [pool.submit(self.obtain_user_info_thread, session, group_list) for session in session_list]
            for future in as_completed(self.future_list):
                error = future.exception(5)
                if error:
                    print(error)
                    self.ms.print_log.emit("任务出错，当前任务停止")
        pool.shutdown()
        self.ms.print_log.emit("-----获取群组成员信息任务完成-----")
        self.ms.show_message_box.emit("提示", "获取群组成员信息完成\n群组成员信息保存在" + self.user_info_path + "目录下")

    # 开始获取群组成员信息
    def start_obtain(self):
        self.ui.tabWidget.setCurrentIndex(1)
        # 写出配置文件
        self.set_values("conf.ini")
        # 初始化变量
        self.session_path = ""
        self.user_info_path = ""
        self.proxy_link = ""
        self.thread_count = 0
        self.group_count = 0
        self.group_file_path = ""
        self.all_group_list = []
        self.group_list_index = 0
        self.process_bar_value = 0
        self.login_time = 0

        # 1.检查所需的各个信息是否填写完整
        # 检查群组文件是否选择
        self.group_file_path = self.ui.lineEdit_group.text().strip()
        if not self.group_file_path:
            self.ms.show_message_box.emit("警告", "未选择群组文件!")
            return
        self.ms.print_log.emit("选择的群组文件为:" + self.group_file_path)

        # 获取group_list
        group_list = []
        try:
            group_list = self.get_group_list(self.group_file_path)
            self.all_group_list = group_list
        except:
            self.ms.print_log.emit("读取群组文件失败，请检查后重新尝试")
            self.ms.show_message_box.emit("警告", "读取群组文件失败!")
            return
        # 显示导入的群组数
        self.ms.show_info.emit(self.ui.label_import_groups, str(len(group_list)))

        # 获取session文件
        self.session_path = self.ui.lineEdit_sessions.text().strip()
        if not self.session_path:
            self.ms.show_message_box.emit("警告", "未选择sessions文件夹!")
            return
        self.ms.print_log.emit("选择的session文件为:" + self.session_path)

        # 获取用户信息保存文件夹
        self.user_info_path = self.ui.lineEdit_user_info.text().strip()
        if not self.user_info_path:
            self.ms.show_message_box.emit("警告", "未选择用户信息保存文件夹!")
            return
        self.ms.print_log.emit("选择的用户信息保存文件夹为:" + self.user_info_path)

        # 获取session_list
        session_list = []
        try:
            session_list = self.get_session_list(self.session_path)
        except:
            self.ms.print_log.emit("读取session文件失败，请检查后重新尝试")
            return

        # 显示导入的session数
        self.ms.show_info.emit(self.ui.label_import_users, str(len(session_list)))

        # 获取保存格式
        if self.ui.radioButton_txt.isChecked():
            self.save_type = "txt"
        if self.ui.radioButton_excel.isChecked():
            self.save_type = "excel"

        # 检查代理IP链接是否填写
        self.proxy_link = self.ui.lineEdit_ip.text().strip()
        if not self.proxy_link:
            self.ms.show_message_box.emit("警告", "未指定代理ip链接!")
            return

        # 获取线程数
        self.thread_count = int(self.ui.comboBox_thread.currentText().strip())

        # 获取获取群组总数
        self.group_count = len(group_list)

        # 获取指定的登录时间
        self.login_time = int(self.ui.comboBox_login_time.currentText().strip())

        # 判断最终的获取群组数
        if (len(group_list) <= self.group_count) and (len(group_list) <= len(session_list) * self.obtain_group_count):
            self.group_count = len(group_list)
        if ((len(session_list) * self.obtain_group_count) <= len(group_list)) and ((len(session_list) * self.obtain_group_count) <= self.group_count):
            self.group_count = len(session_list) * self.obtain_group_count

        # 根据获取群组总数设置新的session_list
        new_session_list = []
        session_count = 0
        if (self.group_count % self.obtain_group_count) == 0:
            session_count = int(self.group_count / self.obtain_group_count)
        else:
            session_count = int(self.group_count / self.obtain_group_count) + 1

        if session_count > len(session_list):
            session_count = len(session_list)
        new_session_list = session_list[0:session_count]

        # 根据session总数设置新的获取群组列表
        new_group_list = []
        new_group_list = group_list[0:self.group_count]

        # 显示可获取群组数
        self.ms.show_info.emit(self.ui.label_obtainable_groups, str(self.group_count))

        # 设置进度条
        self.ui.progressBar.setMaximum(self.group_count)
        self.ui.progressBar.setValue(0)
        self.ui.progressBar.setFormat("(获取中)%v/%m(群组总数)")

        thread = Thread(target=self.start_obtain_thread, args=(new_group_list, new_session_list, ))
        thread.setDaemon(True)
        thread.start()

        # 2.清空程序运行日志
        # 3.开始发送

    # 停止发送消息线程
    def stop_obtain_thread(self):
        for future in self.future_list:
            future.cancel()
        self.ms.show_message_box.emit("提示", "群组成员信息获取会在当前账号获取完成后停止")

    # 停止发送消息
    def stop_obtain(self):
        thread = Thread(target=self.stop_obtain_thread)
        thread.setDaemon(True)
        thread.start()

