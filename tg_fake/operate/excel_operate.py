import xlrd
import xlwt
import openpyxl


class MyExcel:
    def __init__(self, file_path):
        self.file_path = file_path

    # 从xls文件获取群组链接
    def get_url_xls(self):
        work_book = xlrd.open_workbook(self.file_path)
        sheet = work_book.sheet_by_index(0)
        group_url_list = []
        col_num = sheet.ncols
        row_num = sheet.nrows
        for col in range(col_num):
            if sheet.cell_value(0, col) == "群链接":
                for row in range(1, row_num):
                    group_url_list.append(sheet.cell_value(row, col))
        return group_url_list

    # 从xlsx文件获取群组链接
    def get_url_xlsx(self):
        work_book = openpyxl.load_workbook(self.file_path)
        sheet = work_book.worksheets[0]
        group_url_list = []
        col_num = sheet.max_column
        row_num = sheet.max_row
        for col in range(1, col_num + 1):
            if sheet.cell(1, col).value == "群链接":
                for row in range(2, row_num + 1):
                    group_url_list.append(sheet.cell(row, col).value)
        work_book.close()
        return group_url_list

    # 删除xls某些行
    def rewrite_xls(self, index, save_path):
        work_book = xlrd.open_workbook(self.file_path)
        sheet = work_book.sheet_by_index(0)
        rows = sheet.nrows
        cols = sheet.ncols
        work_book_new = xlwt.Workbook()
        sheet_new = work_book_new.add_sheet("userinfo")
        for col in range(cols):
            sheet_new.write(0, col, sheet.cell_value(0, col))

        count = 1
        for row in range(index, rows):
            for col in range(cols):
                sheet_new.write(count, col, sheet.cell_value(row, col))
            count += 1

        work_book_new.save(save_path)


if __name__ == '__main__':
    my_excel = MyExcel("../1.xlsx")
    list = my_excel.get_url_xlsx()
    print(list)